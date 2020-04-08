# Kleine Hochzeitsgästeliste
# Für die Hochzeit von Theresa Markgraf und David Kirner
# Copyright 2020
# by Mark Fechner
 
# Probleme zu beheben: Einlesen der Excel
#                      Integration der Tischverteilung
 
import openpyxl
import csv
from tkinter import *
 
gaeste = []
 
softwareRunning = True
EingabeWiederholung = True
anzahlGaeste = 0
Tischgaeste = 0
Kontrolle = [0,1,2,3,4,5,6,7,8,9,10]
Kontrollbuchstaben = ["a","A","b","B","c","C","d","D","e","E","f","F","g","G","h","H","i","I","j","J","k","K","l","L","m","M","n","N","o","O","p","P","q","Q","r","R","s","S","t","T","u","U","v","V","w","W","x","X","y","Y","z","Z"]
 
TischEins = {"LinksOben": " ", "RechtsOben": " ", "LinksUnten": " ", "RechtsUnten": " "}
TischZwei = {"LinksOben": " ", "RechtsOben": " ", "LinksUnten": " ", "RechtsUnten": " "}
TischDrei = {"LinksOben": " ", "RechtsOben": " ", "LinksUnten": " ", "RechtsUnten": " "}
TischVier = {"LinksOben": " ", "RechtsOben": " ", "LinksUnten": " ", "RechtsUnten": " "}
TischFunf = {"LinksOben": " ", "RechtsOben": " ", "LinksUnten": " ", "RechtsUnten": " "}
 
Tische = [TischEins, TischZwei, TischDrei, TischVier, TischFunf]
 
def TischOrdnung(Ordnung):
    print(TischEins["LinksOben"] + "|" + TischEins["RechtsOben"])
    print(TischEins["LinksUnten"] + "|" + TischEins["RechtsUnten"])
    print("-+-+-")
    print(TischZwei["LinksOben"] + "|" + TischZwei["RechtsOben"])
    print(TischZwei["LinksUnten"] + "|" + TischZwei["RechtsUnten"])
    print("-+-+-")
    print(TischDrei["LinksOben"] + "|" + TischDrei["RechtsOben"])
    print(TischDrei["LinksUnten"] + "|" + TischDrei["RechtsUnten"])
    print("-+-+-")
    print(TischVier["LinksOben"] + "|" + TischVier["RechtsOben"])
    print(TischVier["LinksUnten"] + "|" + TischVier["RechtsUnten"])
    print("-+-+-")
    print(TischFunf["LinksOben"] + "|" + TischFunf["RechtsOben"])
    print(TischFunf["LinksUnten"] + "|" + TischFunf["RechtsUnten"])
    print("-+-+-")
 
def Gästeliste_Import():
    #wb = openpyxl.load_workbook("TÄST01.xlsx")
    #print(wb.sheetnames)
    wb = openpyxl.load_workbook(input("Bitte den Namen der zu importierenden Excel eingeben: ") + '.xlsx')
    sheet = wb.active
    for row in sheet.rows:
        #print(row[0].value)
        gaeste_import = []
        #gaeste_import.append(row[0].value)
        gaeste.append(row[0].value)
        #gaeste_import = [row[0].value]
        #print(gaeste_import)
 
def Gästeliste_Export():
    wb = openpyxl.load_workbook('example.xlsx')
    sheet = wb.active
    sheet.title = 'Gästesheet'
    for i in range(0,len(gaeste)):
        cellref = sheet.cell(row=i + 1, column=1)
        cellref.value = gaeste[i]
    Dateiname = input("Bitte Namen für Speicherdatei eingeben: ")
    wb.save(Dateiname + '.xlsx')

print("MMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMWk;oNMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMM")
print("MMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMWx..'OMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMM")
print("MMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMWx'...dWMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMM")
print("MMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMM0,....cNMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMM")
print("NKKKXNWMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMNko;...:XMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMM")
print("Xd,',;:loxOKNWMWMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMWKOK0x:.;KMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMM")
print("MWO:.......l0KKKXNWMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMNO0KKK0lcKMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMM")
print("MMMNk:.....cKKK0000KXNWMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMWKOKKKKK0kXMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMM")
print("MMMMMNkc...;OKKKKKK0000KXNWMMMMMMMMMMMMMMMMMMMMMMMMMMN0OKKKKK0ONMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMM")
print("MMMMMMMW0o;.lO00KKKKKKK000KXWMMMMMMMMMMMMMMMMMMMMMMMMXO0KKKKKkOWMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMM")
print("MMMMMMMMMWNkdxkkOO00KKKKKK000KNWMWWWWNWWWWMMMMMMMMMMMXOKKKKK0k0MMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMM")
print("MMMMMMMMMMMMWNK0kxxkO00KKKKKK0O0KK0000000KKKKXXNWMMMMKOKKKKKOkKMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMM")
print("MMMMMMMMMMMMMMMMWX0OOkkkO00KKK00KKKKKKKKKKKK00000KXXWKOKKKK0kkNMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMM")
print("MMMMMMMMMMMMMMMMMMMMWNK00kxOKKKKKKKKKKKKKKKKKKKKK0OO0kk0KKKOx0MMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMM")
print("MMMMMMMMMMMMMMMMWMMMMMMMN0OKKKKKKKKKKKKKKKKKKKKKKK0KK0OO0K0kkNMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMM")
print("MMMMMMMMMMMMMMMMMMMMMMMN0OKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKK0kxKMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMM")
print("MMMMMMMMMMMMMMMMMMMMMMW0OKKklokOOKKKKKKKKKKKKKKKKKKKKKKKKOx0WMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMM")
print("MMMMMMMMMMMMMMMMMMMMMMXO0Kk,.:Ooo0KKKKKKKKKKKKKKKKKKKKKKK0ONMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMM")
print("MMMMMMMMMMMMMMMMMMMMMNOOKKO:...;xKKKKKKKKKKKKKK0kO0kx0KKK0OXMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMM")
print("MMMMMMMMMMMMMMMMMMMMXxlokKK0kxk0KKKKKKOk0KKKKKKo:xKx,:OKKO0NMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMM")
print("MMMMMMMMMMMMMMMMMMMXxcccoOKKKKK0OOOOkxdx0KKKKKKo..'..;kKKOKWMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMM")
print("MMMMMMMMMMMMMMMMMMM0lccclkKKKKKKx:,;,''':okO00K0xcccok0K0OXMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMM")
print("MMMMMMMMMMMMMMMMMMM0lclcoOKKKKKKk:;cccc;,'';d0KKKKKKKKKK0ONMMWWNXXXKKNMMMMMMMMMMMMMMMMMMMMMMMMMMMMMM")
print("MMMMMMMMMMMMMMMMMMMNxccoOKKKKKKKOooddddol:,l0KKKKKKKKOxdoxKXK000000OkONWMMMMMMMMMMMMMMMMMMMMMMMMMMMM")
print("MMMMMMMMMMMMMMMMMMMMKxk0KKKKKKKK0doddddddod0KKKKKKK0xlcccoO00KKKKKKK0kONMMMMMMMMMMMMMMMMMMMMMMMMMMMM")
print("MMMMMMMMMMMMMMMMMWWMWXO0KKKKKKKKKxodddddox0KKKKKK0OdcccccxKKKKKKKKKKOkXWMMMMMMMMMMMMMMMMMMMMMMMMMMMM")
print("MMMMMMMMMMMMMMMMWWNXXKOkO0000KKKKOdoddodkKKKKKKKKOxlccccx0KKKKKKKKOkxOKXXXXXXXXXNNWWMMMMMMMMMMMMMMMM")
print("MMMMMMMMMMMMMMNX00000000000OOOOO0K0kxkO0KKKKKKKKK00xlcok0KKKK0KK0OkO000000000000000KKKKKXXNNWMMMMMMM")
print("MMMMMMMMMMMMWXOO0KKKKKKKKKKKKK0OOOO0KKKKKKKKKKKK0OOOkk0KKKKKKK0Oxk0KKKKKKKKKKKKKKKKKKKK000000KKKXNWW")
print("MMMMMMMMMMMMXO0KKKKKKKKKKKKKKKKKK0kxO0000KKKKK0OOO00KKKKKKKK0Okxx0KKKKKKKKKKKKKKKKKKKKKKKKKKKK000OON")
print("MMMMMMMMMMMMXO0KKKKKKKKKKKKKKKKKK0kxOOO00KKKKKKKKKKKKKKKK0Okkk00OK0OO00KKKKKKKKKKKKKKKKKKKKKKKK000XW")
print("MMMMMMMMMMMMN0O0K00K0KKKKKKKK000OkxOKKKKKKKKKKKKKKKKKK0OOkxk0NN00KOkkkkOO00KKKKKKKKKKKKKKKKKKK0OXWMM")
print("MMMMMMMMMMMMMNKkkkkOOOOOOOOkkkkkOO0KKKKKKKKKKKKKKKKK00kxxk0NWMXO00OkkkkkkkkkO00KKKKKKKKKKKKK0OKNMMMM")
print("MMMMMMMMMMMMMMWNK0OkxxxxxkkkO0KKKKKKKKKKKKKKKKKKKKK0OxkOKNWMMMKOK0kkkkxxkkkkkkkOO0KKKKKKKK0O0NMMMMMM")
print("MMMMMMMMMMMMMMMMWWNOk0000KKKKKKKKKKKKKKKKKKKKKKKKKK0kx0WWMMMMW0OKOkkkxk000OkkxxxxkkO0KKK000XWMMMMMMM")
print("MMMMMMMMMMMMMMMMMMKO0KKKKKKKKKKKKKKKKKKKKKKKKKKKKKKK0OXWX0KXXKO0KOkkxONMMMMWNXXKK00OOOOO0XWMMMMMMMMM")
print("MMMMMMMMMMMMMMMMMW0OKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKK0OXWOxkOO0000OkxONMMMMMMMMMMMMMMWNXXWMMMMMMMMMMM")
print("MMMMMMMMMMMMMMMMMKO0KKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKK0OKMKxxkkO000kxONMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMM")
print("MMMMMMMMMMMMMMMMN00KKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKO0WNkdxxxkOkxONMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMM")
print("MMMMMMMMMMMMMMMWKOKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKK0O0Xx:::oKWNNWMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMM")
print("MMMMMMMMMMMMMMMXO0KKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKkc::;,,cKWMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMM")
print("MMMMMMMMMMMMMMW0OKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKxcc:;,lXMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMM")
print("MMMMMMMMMMMMMMXO0KKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKK0O0XKO0NMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMM")
print("MMMMMMMMMMMMMWKOKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKO0NMWMMWMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMM")
print("MMMMMMMMMMMMMW0OKKKKKKKKKKKKKKKKKKKK000000000000000KKKKK0OXWWWMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMM")
print("MMMMMMMMMMMMMMXOOOkOOOO00000000OOOOOkkkkkkkkkkkkkkkkOO0KO0WMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMM")
print("MMMMMMMMMMMMMMWXOxxxkkkkkkkkxxxxxxddddddxkkkkkkkkkkkkkkxxKNWMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMM")
print("MMMMMMMMMMMMWNK00Okkxkkxddk000KXXXKKKK00000OOOOxdxxxxxkkkOOkOKNWMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMM")
print("MMMMMMMMMMMXOxdxOOOkkkOOOKWMWMMMMMMMMMMMMMMMMWWX0OOOO00000OkxddONMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMM")
print("MMMMMMMMMWM0ddxkkO0KKXNWWMMMMMMMMMMMMMMMMMMMMMMMMMMWWNXXKK0OOkdxXMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMMM") 
print("Herzlich Willkommen beim Gästetool für die Hochzeit von Theresa und David!")
print("Mit diesem kleinen Tool können die Gäste erfasst, gezählt und die Anzahl der nötigen Tische ermittelt werden.\n")
Tischgaeste = input("Wie viele Personen sollen an einen Tisch gesetzt werden? \n : ")
Tischgaeste = int(Tischgaeste)
if Tischgaeste not in Kontrolle:
    print("Das war wohl verkehrt!")
    Tischgaeste = input("Nächster Versuch...? \n :")
    Tischgaeste = int(Tischgaeste)
 
 
while softwareRunning == True: # Schleife für das Hauptmenü
    auswahl = input("Bitte auswählen:\n 1 Gast hinzufügen,\n 2 Gast aus Liste entfernen,\n 3 Gästeliste ausgeben,\n 4 Gästeexcel speichern,\n 5 Gästeexcel importieren,\n 6 Tischordnung darstellen,\n 7 Programm beenden. \n")
    auswahl = int(auswahl)
    if auswahl == 1:
        while EingabeWiederholung == True: # Schleife für die ununterbrochene Eingabe von Gästen
            anzahlGaeste += 1
            tische = int(int(anzahlGaeste) / int(Tischgaeste))
            gast = input(str("Bitte den Namen des Gasts eingeben: (Oder e für Ende der Eingabe) "))
            gast = str(gast + ", ")
            if gast[0] not in Kontrollbuchstaben:
                gast = input(str("Bitte einen Namen eingeben...! "))
            gaeste.append(gast)
            
            if gast == "e, ":
                gaeste.remove("e, ")
                anzahlGaeste = anzahlGaeste -1
                EingabeWiederholung = False # Ender der Schleife
                
    elif auswahl == 2:
        anzahlGaeste = anzahlGaeste -1
        tische = int(int(anzahlGaeste) / int(Tischgaeste))
        zuEntfernen = input("Bitte den Namen eingeben, der gelöscht werden soll: ")
        try:
            gaeste.remove(zuEntfernen)
            print(zuEntfernen + " wurde erfolgreich aus der Liste entfernt.")
        except:
            if zuEntfernen is not gaeste:
                auswahl = 2
 
    elif auswahl == 3:
        print(gaeste)
        print("Das sind " + str(anzahlGaeste) + " Gäste bisher.")
        print("Damit werden " + str(tische) + " Tische benötigt.\n")
 
    elif auswahl == 4:
        Gästeliste_Export()
        
    elif auswahl == 5:
        Gästeliste_Import()
 
    elif auswahl == 6:
        TischOrdnung(Tische)
 
    elif auswahl == 7:
        print("Vielen Dank für die Nutzung dieser Software!")
        softwareRunning = False
        
    else:
        print("Überfordert schon die Eingabe einer einzelnen Zahl???")
 
 
 

